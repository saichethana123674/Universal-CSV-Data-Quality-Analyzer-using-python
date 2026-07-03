from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import datetime
import os


def generate_excel_report(df, score, dataset_name):

    os.makedirs("reports", exist_ok=True)

    wb = Workbook()

    ws = wb.active

    ws.title = "Data Quality Report"

    # ==========================
    # Title
    # ==========================

    ws["A1"] = "UNIVERSAL CSV DATA QUALITY REPORT"
    ws["A1"].font = Font(size=18, bold=True)

    # ==========================
    # Dataset Information
    # ==========================

    ws["A3"] = "Dataset Name"
    ws["B3"] = dataset_name

    ws["A4"] = "Rows"
    ws["B4"] = df.shape[0]

    ws["A5"] = "Columns"
    ws["B5"] = df.shape[1]

    ws["A6"] = "Generated On"
    ws["B6"] = datetime.now().strftime("%d-%m-%Y %H:%M")

    # ==========================
    # Quality Metrics
    # ==========================

    ws["A8"] = "Quality Metrics"
    ws["A8"].font = Font(bold=True)

    ws["A9"] = "Missing Values"
    ws["B9"] = int(df.isnull().sum().sum())

    ws["A10"] = "Duplicate Rows"
    ws["B10"] = int(df.duplicated().sum())

    ws["A11"] = "Empty Columns"
    ws["B11"] = len(df.columns[df.isnull().all()])

    ws["A12"] = "Quality Score"
    ws["B12"] = f"{score:.2f}"

    # ==========================
    # Column Information
    # ==========================

    ws["A14"] = "Column Name"
    ws["B14"] = "Data Type"
    ws["C14"] = "Missing Values"

    ws["A14"].font = Font(bold=True)
    ws["B14"].font = Font(bold=True)
    ws["C14"].font = Font(bold=True)

    row = 15

    for col in df.columns:

        ws.cell(row=row, column=1).value = col
        ws.cell(row=row, column=2).value = str(df[col].dtype)
        ws.cell(row=row, column=3).value = int(df[col].isnull().sum())

        row += 1

    wb.save("reports/Data_Quality_Report.xlsx")

    print("✅ Excel Report Generated Successfully!")